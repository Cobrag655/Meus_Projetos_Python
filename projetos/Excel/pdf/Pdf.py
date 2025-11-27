#importar a bibli
import openpyxl 
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import os

arquivo = "banco.xlsx"

#se o arquivo nao existir cria outro
if not os.path.exists(arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"

    #cabeçario para planilhas
    ws.append(["Nome", "Idade", "Nota"])
    wb.save(arquivo)

#abrindo o arquivo existente
wb = openpyxl.load_workbook(arquivo)
ws = wb["Alunos"]

#add as info (dic)
alunos = [
    ["Gabriell", 16, 8.5],
    ["Pedro", 18, 7.0],
    ["Guilherme", 17, 8.5]
]
for aluno in alunos:
    ws.append(aluno)

#salvando o documento
wb.save(arquivo)

#gerar um pdf a partir do excel
#abrir o arquivo excel
wb = openpyxl.load_workbook("banco.xlsx")
ws = wb.active #pegar a primeira planilhia

#criando pdf
pdf = canvas.Canvas("saida.pdf", pagesize = A4)
largura, altura = A4

#titulo do documento
pdf.setFont("Helvetica", 14)
pdf.drawString(50,altura - 50, "Relatorio do projeto pdf, direto do excel.\n")

#posicao incial do arquivo
y = altura - 100

#ler os dados da planilhia e escrever no pdf
for linha in ws.iter_rows(values_only=True):
    texto = " | ".join([str(celula) if celula is not None else "" for celula in linha])
    pdf.drawString(50, y, texto)
    y -= 20 #descer a linha do pdf
    
    if y < 50: #se a pagina acabar ele cria outra
        pdf.showPage()
        pdf.setFont("Helvetica", 15)
        y = altura - 50


#Salvar pdf
pdf.save()
print("PDF gerado com susses: saida.pdf")
